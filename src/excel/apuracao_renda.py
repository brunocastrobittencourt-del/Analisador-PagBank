from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os


class ApuracaoRenda:

    COLUNAS = ["B", "C", "D", "E", "F", "G"]

    LINHA_MESES = 13
    LINHA_INICIAL_DIAS = 14
    LINHA_TOTAL = 45
    CELULA_MEDIA = "F47"

    def __init__(self, modelo, pasta_resultados):

        self.modelo = modelo
        self.pasta_resultados = pasta_resultados

    def gerar(self, apuracao):

        wb = load_workbook(self.modelo)
        ws = wb.active

        meses = list(apuracao.meses.keys())

        # =====================================
        # COLUNA DOS DIAS
        # =====================================

        for dia in range(1, 32):

            linha = self.LINHA_INICIAL_DIAS + (dia - 1)

            celula = ws[f"A{linha}"]
            celula.value = dia
            celula.font = Font(bold=True)
            celula.alignment = Alignment(horizontal="center")

        # =====================================
        # MESES
        # =====================================

        for indice, mes in enumerate(meses):

            if indice >= 6:
                break

            coluna = self.COLUNAS[indice]

            celula = ws[f"{coluna}{self.LINHA_MESES}"]
            celula.value = mes
            celula.font = Font(bold=True)
            celula.alignment = Alignment(horizontal="center")

        # =====================================
        # VALORES DIÁRIOS
        # =====================================

        for indice, mes in enumerate(meses):

            if indice >= 6:
                break

            coluna = self.COLUNAS[indice]

            dias = apuracao.meses[mes]

            for dia in range(1, 32):

                linha = self.LINHA_INICIAL_DIAS + (dia - 1)

                valor = dias.get(dia, 0)

                if valor > 0:

                    celula = ws[f"{coluna}{linha}"]

                    celula.value = valor
                    celula.number_format = 'R$ #,##0.00'
                    celula.alignment = Alignment(horizontal="center")

        # =====================================
        # TOTAL
        # =====================================

        ws["A45"] = "TOTAL"
        ws["A45"].font = Font(bold=True)
        ws["A45"].alignment = Alignment(horizontal="center")

        for indice, mes in enumerate(meses):

            if indice >= 6:
                break

            coluna = self.COLUNAS[indice]

            celula = ws[f"{coluna}{self.LINHA_TOTAL}"]

            celula.value = apuracao.total_meses[mes]
            celula.number_format = 'R$ #,##0.00'
            celula.font = Font(bold=True)

            celula.fill = PatternFill(
                fill_type="solid",
                fgColor="FFF2CC"
            )

            celula.alignment = Alignment(horizontal="center")

        # =====================================
        # MÉDIA
        # =====================================

        media = ws[self.CELULA_MEDIA]

        media.value = apuracao.media
        media.number_format = 'R$ #,##0.00'
        media.font = Font(
            bold=True,
            size=12
        )

        media.alignment = Alignment(horizontal="center")

        # =====================================
        # LARGURA DAS COLUNAS
        # =====================================

        larguras = {
            "A": 8,
            "B": 16,
            "C": 16,
            "D": 16,
            "E": 16,
            "F": 16,
            "G": 16,
        }

        for coluna, largura in larguras.items():

            ws.column_dimensions[coluna].width = largura

        # =====================================
        # SALVAR
        # =====================================

        caminho = os.path.join(
            self.pasta_resultados,
            "APURACAO_DE_RENDA.xlsx"
        )

        wb.save(caminho)

        return caminho