from openpyxl import Workbook
import os


class GeradorExcel:

    @staticmethod
    def gerar(totais, pasta_resultados):

        wb = Workbook()
        ws = wb.active

        ws.title = "Recebimentos"

        ws["A1"] = "DATA"
        ws["B1"] = "TOTAL RECEBIDO"

        linha = 2

        for data, valor in totais.items():
            ws.cell(row=linha, column=1).value = data
            ws.cell(row=linha, column=2).value = round(valor, 2)
            linha += 1

        caminho = os.path.join(
            pasta_resultados,
            "Recebimentos_Diarios.xlsx"
        )

        wb.save(caminho)

        return caminho


    @staticmethod
    def gerar_movimentacoes(movimentacoes, pasta_resultados):

        wb = Workbook()
        ws = wb.active

        ws.title = "Movimentações"

        ws.append([
            "Data",
            "Tipo",
            "Nome",
            "Valor",
            "Descrição"
        ])

        for mov in movimentacoes:
            ws.append([
                mov.data,
                mov.tipo,
                mov.nome,
                mov.valor,
                mov.descricao
            ])

        caminho = os.path.join(
            pasta_resultados,
            "Movimentacoes_Processadas.xlsx"
        )

        wb.save(caminho)

        return caminho


    @staticmethod
    def gerar_resumo(
        receita_total,
        quantidade_dias,
        media_diaria,
        maior_dia,
        menor_dia,
        pasta_resultados
    ):

        wb = Workbook()
        ws = wb.active

        ws.title = "Resumo Financeiro"

        ws.append(["Indicador", "Valor"])

        ws.append(["Receita Total", receita_total])
        ws.append(["Quantidade de Dias", quantidade_dias])
        ws.append(["Média Diária", media_diaria])
        ws.append(["Maior Dia", maior_dia])
        ws.append(["Menor Dia", menor_dia])

        caminho = os.path.join(
            pasta_resultados,
            "Resumo_Financeiro.xlsx"
        )

        wb.save(caminho)

        return caminho